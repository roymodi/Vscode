import openpyxl as op 

c=op.load_workbook('C:\\Users\\bidyut\\Vscode\\a.xlsx')
print(c.sheetnames)

s=c['my']
x=s.cell(row=1,column=2).value
print(x)